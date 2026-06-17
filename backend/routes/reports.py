from flask import Blueprint, request, jsonify, send_file
from extensions import db
from models import *
from sqlalchemy import func, extract
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import openpyxl
from openpyxl.styles import Font, Alignment
from io import BytesIO
from flask_jwt_extended import jwt_required
from utils import role_required
from routes import reports_bp


@reports_bp.route('/sales-summary', methods=['GET'])
@jwt_required()
@role_required('admin', 'manager', 'staff')
def sales_summary():
    try:
        today = datetime.utcnow().date()
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')

        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        else:
            start_date = today.replace(day=1)

        if end_date_str:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        else:
            end_date = today

        query = Sale.query.filter(Sale.sale_date >= start_date, Sale.sale_date <= end_date)

        sales_count = query.count()

        revenue = db.session.query(func.sum(Sale.total_amount)).filter(
            Sale.sale_date >= start_date, Sale.sale_date <= end_date
        ).scalar() or 0

        discount_given = db.session.query(func.sum(Sale.discount)).filter(
            Sale.sale_date >= start_date, Sale.sale_date <= end_date
        ).scalar() or 0

        tax_collected = db.session.query(func.sum(Sale.tax)).filter(
            Sale.sale_date >= start_date, Sale.sale_date <= end_date
        ).scalar() or 0

        items_sold = db.session.query(func.sum(SaleItem.quantity)).join(Sale).filter(
            Sale.sale_date >= start_date, Sale.sale_date <= end_date
        ).scalar() or 0

        profit_data = db.session.query(
            func.sum((SaleItem.unit_price - Product.purchase_price) * SaleItem.quantity)
        ).join(Sale).join(Product, Product.id == SaleItem.product_id).filter(
            Sale.sale_date >= start_date, Sale.sale_date <= end_date
        ).scalar() or 0

        return jsonify({
            'success': True,
            'data': {
                'sales_count': sales_count,
                'revenue': float(revenue),
                'discount_given': float(discount_given),
                'tax_collected': float(tax_collected),
                'profit': float(profit_data),
                'items_sold': items_sold
            },
            'message': 'Sales summary retrieved successfully'
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'data': None, 'message': str(e)}), 500


@reports_bp.route('/purchase-summary', methods=['GET'])
@jwt_required()
@role_required('admin', 'manager', 'staff')
def purchase_summary():
    try:
        today = datetime.utcnow().date()
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')

        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        else:
            start_date = today.replace(day=1)

        if end_date_str:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        else:
            end_date = today

        purchase_count = Purchase.query.filter(
            Purchase.purchase_date >= start_date, Purchase.purchase_date <= end_date
        ).count()

        total_spent = db.session.query(func.sum(Purchase.total_amount)).filter(
            Purchase.purchase_date >= start_date, Purchase.purchase_date <= end_date
        ).scalar() or 0

        return jsonify({
            'success': True,
            'data': {
                'purchase_count': purchase_count,
                'total_spent': float(total_spent)
            },
            'message': 'Purchase summary retrieved successfully'
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'data': None, 'message': str(e)}), 500


@reports_bp.route('/profit-loss', methods=['GET'])
@jwt_required()
@role_required('admin', 'manager', 'staff')
def profit_loss():
    try:
        today = datetime.utcnow().date()
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')

        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        else:
            start_date = today.replace(day=1)

        if end_date_str:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        else:
            end_date = today

        revenue = db.session.query(func.sum(Sale.total_amount)).filter(
            Sale.sale_date >= start_date, Sale.sale_date <= end_date
        ).scalar() or 0

        cogs_data = db.session.query(
            func.sum(Product.purchase_price * SaleItem.quantity)
        ).join(SaleItem, SaleItem.product_id == Product.id).join(Sale).filter(
            Sale.sale_date >= start_date, Sale.sale_date <= end_date
        ).scalar() or 0

        gross_profit = float(revenue) - float(cogs_data)

        expenses = db.session.query(func.sum(Expense.amount)).filter(
            Expense.expense_date >= start_date, Expense.expense_date <= end_date
        ).scalar() or 0

        net_profit = float(gross_profit) - float(expenses)

        return jsonify({
            'success': True,
            'data': {
                'revenue': float(revenue),
                'cogs': float(cogs_data),
                'gross_profit': float(gross_profit),
                'expenses': float(expenses),
                'net_profit': float(net_profit)
            },
            'message': 'Profit and loss retrieved successfully'
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'data': None, 'message': str(e)}), 500


@reports_bp.route('/top-products', methods=['GET'])
@jwt_required()
@role_required('admin', 'manager', 'staff')
def top_products():
    try:
        today = datetime.utcnow().date()
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        limit = request.args.get('limit', 10, type=int)

        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        else:
            start_date = today.replace(day=1)

        if end_date_str:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        else:
            end_date = today

        results = db.session.query(
            Product.id,
            Product.name,
            Product.sku,
            func.sum(SaleItem.quantity * SaleItem.unit_price).label('total_revenue'),
            func.sum(SaleItem.quantity).label('total_quantity')
        ).join(SaleItem, SaleItem.product_id == Product.id).join(Sale).filter(
            Sale.sale_date >= start_date, Sale.sale_date <= end_date
        ).group_by(Product.id, Product.name, Product.sku).order_by(
            func.sum(SaleItem.quantity * SaleItem.unit_price).desc()
        ).limit(limit).all()

        top_products_list = []
        for row in results:
            top_products_list.append({
                'product_id': row.id,
                'product_name': row.name,
                'sku': row.sku,
                'total_revenue': float(row.total_revenue),
                'total_quantity': int(row.total_quantity)
            })

        return jsonify({
            'success': True,
            'data': top_products_list,
            'message': 'Top products retrieved successfully'
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'data': None, 'message': str(e)}), 500


@reports_bp.route('/stock-valuation', methods=['GET'])
@jwt_required()
@role_required('admin', 'manager', 'staff')
def stock_valuation():
    try:
        products = Product.query.filter_by(is_active=True).all()

        valuation_list = []
        total_valuation = 0.0
        for product in products:
            stock_value = float(product.stock_quantity * (product.purchase_price or 0))
            total_valuation += stock_value
            valuation_list.append({
                'product_id': product.id,
                'product_name': product.name,
                'sku': product.sku,
                'category': product.category.name if product.category else None,
                'brand': product.brand.name if product.brand else None,
                'stock_quantity': product.stock_quantity,
                'purchase_price': float(product.purchase_price) if product.purchase_price else 0.0,
                'stock_value': round(stock_value, 2)
            })

        return jsonify({
            'success': True,
            'data': {
                'products': valuation_list,
                'total_valuation': round(total_valuation, 2)
            },
            'message': 'Stock valuation retrieved successfully'
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'data': None, 'message': str(e)}), 500


@reports_bp.route('/customer-report', methods=['GET'])
@jwt_required()
@role_required('admin', 'manager', 'staff')
def customer_report():
    try:
        customer_id = request.args.get('customer_id', type=int)
        if not customer_id:
            return jsonify({'success': False, 'data': None, 'message': 'customer_id is required'}), 400

        customer = Customer.query.get(customer_id)
        if not customer:
            return jsonify({'success': False, 'data': None, 'message': 'Customer not found'}), 404

        sales = Sale.query.filter_by(customer_id=customer_id).order_by(Sale.sale_date.desc()).all()
        sales_data = []
        for sale in sales:
            s = sale.to_dict()
            s['items_count'] = sale.items.count()
            sales_data.append(s)

        warranties = Warranty.query.filter_by(customer_id=customer_id).order_by(Warranty.start_date.desc()).all()
        warranties_data = [w.to_dict() for w in warranties]

        return jsonify({
            'success': True,
            'data': {
                'customer': customer.to_dict(),
                'total_spent': float(customer.total_purchases or 0),
                'purchase_history': sales_data,
                'loyalty_points': customer.loyalty_points or 0,
                'warranties': warranties_data
            },
            'message': 'Customer report retrieved successfully'
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'data': None, 'message': str(e)}), 500


@reports_bp.route('/export-sales', methods=['GET'])
@jwt_required()
@role_required('admin', 'manager')
def export_sales():
    try:
        today = datetime.utcnow().date()
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        file_format = request.args.get('format', 'xlsx')

        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        else:
            start_date = today.replace(day=1)

        if end_date_str:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        else:
            end_date = today

        sales = Sale.query.filter(
            Sale.sale_date >= start_date, Sale.sale_date <= end_date
        ).order_by(Sale.sale_date.desc()).all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Sales Report'

        headers = ['Invoice #', 'Date', 'Customer', 'Subtotal', 'Discount', 'Tax', 'Total', 'Paid', 'Balance', 'Status', 'Type']
        header_font = Font(bold=True, size=11)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for row_idx, sale in enumerate(sales, 2):
            ws.cell(row=row_idx, column=1, value=sale.invoice_number)
            ws.cell(row=row_idx, column=2, value=sale.sale_date.isoformat() if sale.sale_date else '')
            ws.cell(row=row_idx, column=3, value=sale.customer.name if sale.customer else '')
            ws.cell(row=row_idx, column=4, value=float(sale.subtotal or 0))
            ws.cell(row=row_idx, column=5, value=float(sale.discount or 0))
            ws.cell(row=row_idx, column=6, value=float(sale.tax or 0))
            ws.cell(row=row_idx, column=7, value=float(sale.total_amount or 0))
            ws.cell(row=row_idx, column=8, value=float(sale.amount_paid or 0))
            ws.cell(row=row_idx, column=9, value=float(sale.balance_due or 0))
            ws.cell(row=row_idx, column=10, value=sale.payment_status)
            ws.cell(row=row_idx, column=11, value=sale.sale_type)

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 10
        ws.column_dimensions['K'].width = 12

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'sales_export_{start_date.isoformat()}_{end_date.isoformat()}.{file_format}'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'success': False, 'data': None, 'message': str(e)}), 500


@reports_bp.route('/export-stock', methods=['GET'])
@jwt_required()
@role_required('admin', 'manager')
def export_stock():
    try:
        file_format = request.args.get('format', 'xlsx')

        products = Product.query.filter_by(is_active=True).order_by(Product.name).all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Stock Report'

        headers = ['SKU', 'Product Name', 'Category', 'Brand', 'Supplier', 'Purchase Price', 'Selling Price',
                    'Stock Quantity', 'Low Stock Threshold', 'Stock Value', 'Condition']
        header_font = Font(bold=True, size=11)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for row_idx, product in enumerate(products, 2):
            stock_value = float(product.stock_quantity * (product.purchase_price or 0))
            ws.cell(row=row_idx, column=1, value=product.sku)
            ws.cell(row=row_idx, column=2, value=product.name)
            ws.cell(row=row_idx, column=3, value=product.category.name if product.category else '')
            ws.cell(row=row_idx, column=4, value=product.brand.name if product.brand else '')
            ws.cell(row=row_idx, column=5, value=product.supplier.name if product.supplier else '')
            ws.cell(row=row_idx, column=6, value=float(product.purchase_price or 0))
            ws.cell(row=row_idx, column=7, value=float(product.selling_price or 0))
            ws.cell(row=row_idx, column=8, value=product.stock_quantity)
            ws.cell(row=row_idx, column=9, value=product.low_stock_threshold)
            ws.cell(row=row_idx, column=10, value=round(stock_value, 2))
            ws.cell(row=row_idx, column=11, value=product.condition)

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 20
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 12

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'stock_export_{datetime.utcnow().strftime("%Y%m%d")}.{file_format}'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'success': False, 'data': None, 'message': str(e)}), 500
